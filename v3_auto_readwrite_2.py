import sys
import gspread
import pandas as pd
import openpyxl
from datetime import datetime, timedelta
import numpy as np
import tkinter as tk
import win32clipboard as win32CPB #Windows Clipboard


"""

Version 3.2

Date: 12/28/2023

Features & Updates:
1. Local Excel Sheet
2. Six-day Report
3. State Alarm


"""



#PATTERN_Bucket = [(0,195),(1,199),(2,200),(3,202),(4,203),(5,207),(6,211),(7,213),(8,216),(9,218),(10,220),(11,230),(12,231)]
PATTERN_Bucket = [195,199,200,202,203,207,211,213,216,218,220,228,230,231]
MARK_States = [202,211,231] #Alarm config
ALARM_Collection = []

RootPath='e:/Project_WorkSpace/UniUni/AutoDailyReport/'


def getDate():

    # Parse the date string into a datetime object
    Report_Date = datetime.strptime(DATE_INPUT, "%m%d%Y")

    DateTracking_Start = 1
    DateTracking_End = 6


    # date-month-list
    dml = []
    # date-day-list
    ddl = []

    #Date List including maping list
    datelist = []

    for i in range(DateTracking_Start,DateTracking_End+1):

        date = Report_Date - timedelta(days=i)

        dml.append(str(date.month))

        if(date.day < 10):
            ddl.append('0'+str(date.day))
        else:
            ddl.append(str(date.day))
    

        
        datelist = list(zip(dml,ddl))[::-1]

    
    # print(datelist)
    return datelist


def genMapList():

    #Read local excel to dataframe via pandas 
    DATA_DF = pd.read_excel(RootPath+"V3/AZ Rd Assignment.xlsx",sheet_name=None)



    #Read-Write Subbatch number in each sheet
    """
    Format -> result = {} #Dictionary ['11/17': ('sub1,sub2',[STATECount Result])],..]

    """

    result = []
    
    for date in DATE_LIST:

        result_sub_ele = {'date':'','batch':[],'st_result':[]} #{'date': '11/03','batch':['PHX-YE-20231101','PHSUB-202311010828'], 'result':[0,0,...,0,0]}
    
        #
        ws_date_in = str(date[0])+'-'+str(date[1])
        single_date_result = [0] * len(PATTERN_Bucket)
        batch_tmp = str(DATA_DF[ws_date_in].at[1,'Unnamed: 2']).split(',')


        result_sub_ele['date'] = ws_date_in
        result_sub_ele['batch'] = batch_tmp
        result_sub_ele['st_result'] = single_date_result
  

        result.append(result_sub_ele)


    # print(result)

    return result


def displayBatchNum(d_b_mapList):



    #Display string
    ds = ''

    for day in d_b_mapList:
        for sb in day['batch']:
            ds += sb + ','
    


    print('\n')
    print('***** Copy-Paste below to download the entire order list *****')
    print(ds)
    print('***** Batch Number Display Ends *****')
    print('\n')

    #Windows auto clipboard
    win32CPB.OpenClipboard()
    win32CPB.EmptyClipboard()
    win32CPB.SetClipboardText(ds)
    win32CPB.CloseClipboard()

    print('For Windows user: Batch Number has been automatically sent to your clipboard')
    print('\n')


    return 1


def countStates(emptyTotal):

    #Get downloaded order sheet
    sheet = openpyxl.load_workbook(RootPath+'V3/order_lists.xlsx')['Order List']


    # Define the cell range 
    start_row = 2 #Column Name on 1st Row
    end_row = sheet.max_row
    state_column = 'C'
    batch_column = 'E'
    driver_column = 'H'


    #Check each row in the sheet
    for row in range(start_row,end_row + 1):


        state_value = sheet[f'{state_column}{row}'].value
        batch_value = sheet[f'{batch_column}{row}'].value

        res_state_index = PATTERN_Bucket.index(state_value)

        instance_date = '' #The date of the order in this row

            
        
        for date in emptyTotal:
            
            #Find day by batch number
            res = [value for key,value in date.items() if any(batch == batch_value for batch in date['batch'])]
            
            #Update state_result for each day
            if(res != []):

                instance_date = res[0] #Record the date for 'this' row
                res[2][res_state_index]  += 1
                # print(res)



        #Sus states collector
        if(state_value in MARK_States):
        
            ALARM_Collection.append((instance_date,sheet[f'{driver_column}{row}'].value,state_value,batch_value))


    return emptyTotal


def writeIn():

    #Write-in destination
    result_book = openpyxl.load_workbook(RootPath+'V3/V3_Daily_Report_Template.xlsx') #Get the Daily Report Template
    result_sheet = result_book['Master Form for report'] 

    """
    Write-in row&col range
    """

    #Regular
    result_row_first_3 = list(range(5,len(PATTERN_Bucket)+5))
    result_row_last_3 = list(range(24,len(PATTERN_Bucket)+24)) #End + 1

    result_col = ''

    dateWriteIn_row = 3
    dateWriteIn_col = ''

    #Alarm
    ALARMWriteIn_col_date = 'Q'
    ALARMWriteIn_col_driver = 'R'
    ALARMWriteIn_col_state = 'S'
    ALARMWriteIn_col_batNum = 'T'

    ALARMWriteIn_row_head = 5

    

    #Daily data iterating
    for cnt,val in enumerate(TOTAL):
 

        #Overwrite in 'Daily Report', otherwise loading 'Template'
        if(cnt > 0):
            result_book = openpyxl.load_workbook('V3-Auto-Daily-Report.xlsx')
            result_sheet = result_book['Master Form for report'] 

        if cnt == 0:
            result_col = 'E'
            dateWriteIn_col = 'D'
            dateWriteIn_row = 3
            
        elif cnt == 1:
            result_col = 'I'
            dateWriteIn_col = 'H'
            dateWriteIn_row = 3

        elif cnt == 2:
            result_col = 'M'
            dateWriteIn_col = 'L'
            dateWriteIn_row = 3

        elif cnt == 3:
            result_col = 'E'
            dateWriteIn_col = 'D'
            dateWriteIn_row = 22
            
        elif cnt == 4:
            result_col = 'I'
            dateWriteIn_col = 'H'
            dateWriteIn_row = 22

        elif cnt == 5:
            result_col = 'M'
            dateWriteIn_col = 'L'
            dateWriteIn_row = 22


        
        #Write in date
        result_sheet[dateWriteIn_col+str(dateWriteIn_row)] = val['date']
        result_book.save('V3-Auto-Daily-Report.xlsx')

        #Write in state data
        for idx,st in enumerate(val['st_result']):
            #idx = # of each single state
            
            if(cnt < 3):
                result_sheet[result_col+str(result_row_first_3[idx])] = st
            else:
                result_sheet[result_col+str(result_row_last_3[idx])] = st

            result_book.save('V3-Auto-Daily-Report.xlsx')

        
        print(val['st_result'])


    #ALARM info write in
    for record in ALARM_Collection:

        result_book = openpyxl.load_workbook('V3-Auto-Daily-Report.xlsx')
        result_sheet = result_book['Master Form for report'] 

        result_sheet[ALARMWriteIn_col_date+str(ALARMWriteIn_row_head)] = record[0]
        result_sheet[ALARMWriteIn_col_driver+str(ALARMWriteIn_row_head)] = record[1]
        result_sheet[ALARMWriteIn_col_state+str(ALARMWriteIn_row_head)] = record[2]
        result_sheet[ALARMWriteIn_col_batNum+str(ALARMWriteIn_row_head)] = record[3]

        result_book.save('V3-Auto-Daily-Report.xlsx')

        ALARMWriteIn_row_head+=1




        

    return 1

#Main

# Get a date in the format "MM/DD/YYYY" from the user
DATE_INPUT = input("Enter report-generating date (MMDDYYYY): ")

try:

    
    #--[Get Date]--
    DATE_LIST = getDate()
    # print(getDate())
    
    #--[Generate Date-Sub-batch Mapping List]--
    TOTAL = genMapList()
    # print(genMapList())
    
    
    
    print('\n')
    print('W-E-L-C-O-M-E')
    print('\n')
    date = datetime.strptime(DATE_INPUT, "%m%d%Y")
    print(f'Daily Report Generating Day: {date.month}-{date.day}-{date.year}')
    date_start = date - timedelta(days=11)
    date_end = date - timedelta(days=5)
    print('\n')
    
    #--[Display mapping list]--
    displayBatchNum(TOTAL)
    
    
    #--[Input Break]--
    #Waiting for getting dowload sheet done
    flag = input("Press 'Y' if you finished downloading the order sheet, 'N' for quit program: ")
    
    if(flag == 'N'):
        quit()
    
    #--[Count States For Each Sheet]--
    TOTAL  = countStates(TOTAL)
    # print(TOTAL)
    
    #--[Write In]--
    writeIn()



except Exception as e:
    exc_type, exc_obj, exc_tb = sys.exc_info() 
    print(exc_type,exc_tb.tb_lineno)
    quit()






